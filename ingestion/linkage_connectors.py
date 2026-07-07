"""Dyadic/linkage data connectors for the coupling matrix.

Covers: IMF DOTS (trade), CEPII GeoDist, COW Alliance/MID, UN Migrant Stock.
BIS Banking is handled separately due to SDMX complexity.
"""

import logging
from pathlib import Path

import pandas as pd
import pycountry

from ingestion.base import DataConnector

logger = logging.getLogger(__name__)


_fuzzy_iso3_cache = {}

def _fuzzy_iso3(name: str) -> str | None:
    if not isinstance(name, str):
        return None
    name = name.strip()
    if name in _fuzzy_iso3_cache:
        return _fuzzy_iso3_cache[name]
    try:
        result = pycountry.countries.search_fuzzy(name)
        val = result[0].alpha_3 if result else None
    except LookupError:
        val = None
    _fuzzy_iso3_cache[name] = val
    return val


class IMFDOTSConnector(DataConnector):
    """IMF Direction of Trade Statistics — bilateral trade flows."""

    source_name = "imf_dots"

    def download(self) -> Path:
        self._ensure_dirs()
        cache_path = self.raw_dir / "dots_trade.csv"
        if self._is_cached(cache_path):
            return cache_path

        logger.info("[dots] Fetching bilateral trade from IMF via imfp...")
        import imfp

        try:
            if self.countries.empty:
                end_year = 2024
            else:
                end_year = int(self.countries["gdp_year"].iloc[0])
            start_year = end_year - 2  # Only need recent trade for coupling matrix

            frames = []
            country_list = self.country_codes if self.country_codes else []

            # Use IMTS (International Trade in Goods) — DOT was deprecated
            database_id = "IMTS"

            # Fetch in batches to avoid API limits
            batch_size = 10
            for i in range(0, len(country_list), batch_size):
                batch = country_list[i:i + batch_size]
                try:
                    df = imfp.imf_dataset(
                        database_id=database_id,
                        frequency="A",
                        country=batch,
                        start_year=start_year,
                        end_year=end_year,
                    )
                    if df is not None and not df.empty:
                        frames.append(df)
                        logger.info("[dots] Batch %d/%d: %d rows", i // batch_size + 1,
                                    (len(country_list) + batch_size - 1) // batch_size, len(df))
                except Exception as e:
                    logger.warning("[dots] Batch %d failed: %s", i, e)

            if frames:
                combined = pd.concat(frames, ignore_index=True)
                combined.to_csv(cache_path, index=False)
            else:
                pd.DataFrame().to_csv(cache_path, index=False)
                logger.warning("[dots] No DOTS/IMTS data fetched.")
        except Exception as e:
            logger.error("[dots] Trade data fetch failed: %s", e)
            pd.DataFrame().to_csv(cache_path, index=False)

        return cache_path

    def clean(self, raw_path: Path) -> pd.DataFrame:
        raw = pd.read_csv(raw_path)
        if raw.empty:
            return pd.DataFrame(columns=["iso3_i", "iso3_j", "year", "value", "indicator", "source"])

        # DOTS returns columns like: ref_area, counterpart_area, indicator, time_period, obs_value
        col_map = {}
        for c in raw.columns:
            cl = c.lower().strip()
            if "ref_area" in cl or "reporter" in cl or cl == "country":
                col_map[c] = "reporter"
            elif "counterpart" in cl or "partner" in cl:
                col_map[c] = "partner"
            elif "time" in cl or "year" in cl:
                col_map[c] = "year"
            elif "obs_value" in cl or "value" in cl:
                col_map[c] = "value"
            elif "indicator" in cl:
                col_map[c] = "dot_indicator"
        raw = raw.rename(columns=col_map)

        required = {"reporter", "partner", "year", "value"}
        if not required.issubset(raw.columns):
            logger.error("[dots] Missing columns. Got: %s", list(raw.columns))
            return pd.DataFrame(columns=["iso3_i", "iso3_j", "year", "value", "indicator", "source"])

        raw["year"] = pd.to_numeric(raw["year"], errors="coerce").astype("Int64")
        raw["value"] = pd.to_numeric(raw["value"], errors="coerce")

        result = pd.DataFrame({
            "iso3_i": raw["reporter"].str.strip().str.upper(),
            "iso3_j": raw["partner"].str.strip().str.upper(),
            "year": raw["year"],
            "value": raw["value"],
            "indicator": "bilateral_trade_usd",
            "source": "IMF_DOTS",
        }).dropna(subset=["iso3_i", "iso3_j", "year", "value"])

        # Filter to country universe
        if not self.countries.empty:
            universe = set(self.country_codes)
            result = result[result["iso3_i"].isin(universe) & result["iso3_j"].isin(universe)]

        # Remove self-loops
        result = result[result["iso3_i"] != result["iso3_j"]]

        logger.info("[dots] Cleaned: %d bilateral trade rows", len(result))
        return result

    def validate(self, df: pd.DataFrame) -> bool:
        if df.empty:
            logger.warning("[dots] No trade data.")
            return False
        return True


class CEPIIConnector(DataConnector):
    """CEPII GeoDist database — bilateral distance, contiguity, language, colonial history."""

    source_name = "cepii"

    CEPII_URL = "http://www.cepii.fr/distance/dist_cepii.zip"

    def download(self) -> Path:
        self._ensure_dirs()
        cache_path = self.raw_dir / "dist_cepii.xls"
        if self._is_cached(cache_path):
            return cache_path

        logger.info("[cepii] Downloading GeoDist data...")
        try:
            resp = self._http_get(self.CEPII_URL, timeout=120)
            import zipfile, io
            with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
                xls_names = [n for n in zf.namelist() if n.endswith(".xls") or n.endswith(".xlsx")]
                if xls_names:
                    with zf.open(xls_names[0]) as f:
                        cache_path.write_bytes(f.read())
                else:
                    logger.error("[cepii] No Excel files found in ZIP")
                    cache_path.touch()
        except Exception as e:
            logger.error("[cepii] Download failed: %s. Place dist_cepii.xls in %s", e, self.raw_dir)
            cache_path.touch()
        return cache_path

    def clean(self, raw_path: Path) -> pd.DataFrame:
        if not raw_path.exists() or raw_path.stat().st_size < 100:
            return pd.DataFrame(columns=["iso3_i", "iso3_j", "year", "value", "indicator", "source"])

        raw = pd.read_excel(raw_path)

        # CEPII columns: iso_o (origin), iso_d (destination), dist, contig, comlang_off, colony, etc.
        iso_o = next((c for c in raw.columns if c.lower() in ("iso_o", "iso3_o")), None)
        iso_d = next((c for c in raw.columns if c.lower() in ("iso_d", "iso3_d")), None)

        if not iso_o or not iso_d:
            logger.error("[cepii] Cannot find ISO columns in: %s", list(raw.columns))
            return pd.DataFrame(columns=["iso3_i", "iso3_j", "year", "value", "indicator", "source"])

        indicators = {
            "dist": "distance_km",
            "contig": "contiguity",
            "comlang_off": "common_official_language",
            "colony": "colonial_relationship",
            "distcap": "distance_capital_km",
        }

        frames = []
        for col, ind_name in indicators.items():
            if col in raw.columns:
                sub = pd.DataFrame({
                    "iso3_i": raw[iso_o].astype(str).str.strip().str.upper(),
                    "iso3_j": raw[iso_d].astype(str).str.strip().str.upper(),
                    "year": 2020,  # Static data, assign arbitrary year
                    "value": pd.to_numeric(raw[col], errors="coerce"),
                    "indicator": ind_name,
                    "source": "CEPII",
                })
                frames.append(sub.dropna(subset=["value"]))

        result = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(
            columns=["iso3_i", "iso3_j", "year", "value", "indicator", "source"])

        if not self.countries.empty:
            universe = set(self.country_codes)
            result = result[result["iso3_i"].isin(universe) & result["iso3_j"].isin(universe)]

        logger.info("[cepii] Cleaned: %d dyadic rows", len(result))
        return result

    def validate(self, df: pd.DataFrame) -> bool:
        return not df.empty


class COWAllianceConnector(DataConnector):
    """Correlates of War Alliance v4.1 + MID data."""

    source_name = "cow_alliance"

    def download(self) -> Path:
        self._ensure_dirs()
        cache_path = self.raw_dir / "cow_alliance.csv"
        if self._is_cached(cache_path):
            return cache_path
        logger.info("[cow] COW data requires manual download from correlatesofwar.org")
        logger.info("[cow] Place alliance and MID CSVs in %s", self.raw_dir)
        cache_path.touch()
        return cache_path

    def clean(self, raw_path: Path) -> pd.DataFrame:
        frames = []

        # Look for alliance files
        alliance_files = list(self.raw_dir.parent.parent.glob("**/alliance*.csv")) + \
                          list(self.raw_dir.glob("*.csv"))
        for af in alliance_files:
            if af.stat().st_size < 100:
                continue
            try:
                df = pd.read_csv(af)
                # COW alliance format: ccode1, ccode2, defense, neutrality, nonaggression, entente
                if "ccode1" in df.columns and "ccode2" in df.columns:
                    for col, ind in [("defense", "alliance_defense"), ("neutrality", "alliance_neutrality"),
                                     ("entente", "alliance_entente")]:
                        if col in df.columns:
                            sub = df[df[col] == 1][["ccode1", "ccode2"]].copy()
                            sub["iso3_i"] = sub["ccode1"].apply(self._cow_to_iso3)
                            sub["iso3_j"] = sub["ccode2"].apply(self._cow_to_iso3)
                            sub["value"] = 1.0
                            sub["indicator"] = ind
                            sub["source"] = "COW"
                            sub["year"] = 2020
                            frames.append(sub[["iso3_i", "iso3_j", "year", "value", "indicator", "source"]].dropna())
            except Exception as e:
                logger.warning("[cow] Failed to parse %s: %s", af, e)

        result = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(
            columns=["iso3_i", "iso3_j", "year", "value", "indicator", "source"])

        if not self.countries.empty:
            universe = set(self.country_codes)
            result = result[result["iso3_i"].isin(universe) & result["iso3_j"].isin(universe)]

        logger.info("[cow] Cleaned: %d alliance/MID rows", len(result))
        return result

    @staticmethod
    def _cow_to_iso3(ccode: int) -> str | None:
        """Convert COW numeric country code to ISO-3166 alpha-3."""
        COW_MAP = {2: "USA", 20: "CAN", 40: "CUB", 70: "MEX", 200: "GBR", 210: "NLD", 211: "BEL",
                   212: "LUX", 220: "FRA", 225: "CHE", 230: "ESP", 235: "PRT", 255: "DEU",
                   290: "POL", 310: "HUN", 316: "CZE", 317: "SVK", 325: "ITA", 338: "MLT",
                   339: "ALB", 341: "MNE", 343: "MKD", 344: "HRV", 345: "SRB", 346: "BIH",
                   349: "SVN", 350: "GRC", 352: "CYP", 355: "BGR", 360: "ROU", 365: "RUS",
                   366: "EST", 367: "LVA", 368: "LTU", 369: "UKR", 370: "BLR", 371: "ARM",
                   372: "GEO", 373: "AZE", 375: "FIN", 380: "SWE", 385: "NOR", 390: "DNK",
                   395: "ISL", 404: "GIN", 432: "MLI", 433: "SEN", 434: "BEN", 435: "MRT",
                   436: "NER", 437: "CIV", 438: "GNB", 439: "BFA", 450: "LBR", 451: "SLE",
                   452: "GHA", 461: "TGO", 471: "CMR", 475: "NGA", 481: "GAB", 482: "CAF",
                   483: "TCD", 484: "COG", 490: "COD", 500: "UGA", 501: "KEN", 510: "TZA",
                   516: "BDI", 517: "RWA", 520: "SOM", 522: "DJI", 530: "ETH", 531: "ERI",
                   540: "AGO", 541: "MOZ", 551: "ZMB", 552: "ZWE", 553: "MWI", 560: "ZAF",
                   565: "NAM", 570: "LSO", 571: "BWA", 572: "SWZ", 580: "MDG", 581: "COM",
                   590: "MUS", 600: "MAR", 615: "DZA", 616: "TUN", 620: "LBY", 625: "SDN",
                   626: "SSD", 630: "IRN", 640: "TUR", 645: "IRQ", 651: "EGY", 652: "SYR",
                   660: "LBN", 663: "JOR", 666: "ISR", 670: "SAU", 678: "YEM", 679: "YEM",
                   690: "KWT", 692: "BHR", 694: "QAT", 696: "ARE", 698: "OMN", 700: "AFG",
                   701: "TKM", 702: "TJK", 703: "KGZ", 704: "UZB", 710: "CHN", 712: "MNG",
                   713: "TWN", 730: "KOR", 731: "PRK", 740: "JPN", 750: "IND", 770: "PAK",
                   771: "BGD", 775: "MMR", 780: "LKA", 790: "NPL", 800: "THA", 811: "KHM",
                   812: "LAO", 816: "VNM", 820: "MYS", 830: "SGP", 840: "PHL", 850: "IDN",
                   900: "AUS", 920: "NZL"}
        return COW_MAP.get(int(ccode)) if pd.notna(ccode) else None

    def validate(self, df: pd.DataFrame) -> bool:
        return True  # May be empty if data not downloaded


class UNMigrantConnector(DataConnector):
    """UN DESA bilateral migrant stock (bulk Excel, 2024 edition)."""

    source_name = "un_migrant"

    # Direct download URL for the 2024 destination×origin matrix
    MIGRANT_URL = (
        "https://www.un.org/development/desa/pd/sites/"
        "www.un.org.development.desa.pd/files/"
        "undesa_pd_2024_ims_stock_by_sex_destination_and_origin.xlsx"
    )
    # Sheet name and layout (0-indexed column positions in row 11 header):
    #   col 0: Index, col 1: Destination name, col 4: Destination loc-code,
    #   col 5: Origin name, col 6: Origin loc-code,
    #   cols 7-14: years 1990 1995 2000 2005 2010 2015 2020 2024
    DATA_SHEET = "Table 1"
    HEADER_ROW = 11   # 1-indexed
    YEAR_COLS  = [7, 8, 9, 10, 11, 12, 13, 14]  # 0-indexed
    YEAR_VALS  = [1990, 1995, 2000, 2005, 2010, 2015, 2020, 2024]

    def download(self) -> Path:
        self._ensure_dirs()
        cache_path = self.raw_dir / "un_migrant.xlsx"
        if self._is_cached(cache_path) and cache_path.stat().st_size > 100_000:
            return cache_path

        logger.info("[migrant] Downloading UN bilateral migrant stock 2024...")
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet,*/*"
            ),
            "Referer": "https://www.un.org/development/desa/pd/content/international-migrant-stock",
        }
        try:
            resp = self._http_get(self.MIGRANT_URL, timeout=120, headers=headers)
            if resp.status_code == 200 and len(resp.content) > 100_000:
                cache_path.write_bytes(resp.content)
                logger.info("[migrant] Downloaded %d bytes", len(resp.content))
            else:
                logger.warning(
                    "[migrant] Download failed (status=%s, size=%d) — "
                    "place xlsx manually in %s",
                    resp.status_code, len(resp.content), self.raw_dir,
                )
                cache_path.touch()
        except Exception as e:
            logger.warning("[migrant] Download error: %s", e)
            cache_path.touch()

        return cache_path

    def clean(self, raw_path: Path) -> pd.DataFrame:
        _empty = pd.DataFrame(
            columns=["iso3_i", "iso3_j", "year", "value", "indicator", "source"]
        )
        if raw_path.stat().st_size < 100_000:
            # Also scan directory for any manually placed xlsx
            candidates = sorted(
                list(self.raw_dir.glob("*.xlsx")) +
                list(self.raw_dir.parent.parent.glob("**/undesa_pd*.xlsx"))
            )
            if not candidates:
                logger.warning("[migrant] No Excel file found in %s", self.raw_dir)
                return _empty
            raw_path = candidates[-1]  # use most recent
            logger.info("[migrant] Using fallback file: %s", raw_path)

        logger.info("[migrant] Parsing bilateral migrant stock from %s", raw_path.name)
        try:
            import openpyxl
        except ImportError:
            logger.warning("[migrant] openpyxl not installed; run: pip install openpyxl")
            return _empty

        try:
            wb = openpyxl.load_workbook(str(raw_path), read_only=True, data_only=True)
        except Exception as e:
            logger.warning("[migrant] Failed to open workbook: %s", e)
            return _empty

        sheet = self.DATA_SHEET if self.DATA_SHEET in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]

        rows_iter = ws.iter_rows(values_only=True)
        # Skip to header row (1-indexed)
        for _ in range(self.HEADER_ROW - 1):
            next(rows_iter, None)

        records = []
        for row in rows_iter:
            if row[1] is None:
                continue  # skip empty / subtotal rows
            dest_code = row[4]   # UN location code for destination
            orig_code = row[6]   # UN location code for origin
            if dest_code is None or orig_code is None:
                continue
            try:
                dest_code = int(dest_code)
                orig_code = int(orig_code)
            except (TypeError, ValueError):
                continue
            for col_idx, year in zip(self.YEAR_COLS, self.YEAR_VALS):
                try:
                    val = float(row[col_idx])
                except (TypeError, ValueError):
                    continue
                records.append((dest_code, orig_code, year, val))

        if not records:
            logger.warning("[migrant] No records parsed from workbook")
            return _empty

        df = pd.DataFrame(records, columns=["dest_code", "orig_code", "year", "migrant_stock"])
        logger.info("[migrant] Parsed %d raw records", len(df))

        # Map UN numeric location codes → ISO3 via pycountry
        def un_code_to_iso3(code: int) -> str | None:
            # Try numeric ISO 3166-1 code first
            try:
                c = pycountry.countries.get(numeric=str(int(code)).zfill(3))
                return c.alpha_3 if c else None
            except Exception:
                return None

        df["iso3_i"] = df["dest_code"].map(un_code_to_iso3)  # destination = receiving country
        df["iso3_j"] = df["orig_code"].map(un_code_to_iso3)  # origin = sending country

        # Keep only country-to-country rows (regions like 900, 1833 etc. return None)
        df = df.dropna(subset=["iso3_i", "iso3_j"])
        df = df[df["iso3_i"] != df["iso3_j"]]  # no self-migration
        df = df[df["migrant_stock"] > 0]

        df["value"] = df["migrant_stock"]
        df["indicator"] = "bilateral_migrant_stock"
        df["source"] = "un_migrant_2024"

        result = df[["iso3_i", "iso3_j", "year", "value", "indicator", "source"]]
        logger.info(
            "[migrant] Cleaned: %d bilateral pairs, %d unique destinations, years %s",
            len(result), result["iso3_i"].nunique(), sorted(result["year"].unique()),
        )
        return result

    def validate(self, df: pd.DataFrame) -> bool:
        return True
